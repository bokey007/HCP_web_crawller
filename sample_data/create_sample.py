"""Generate a sample HCP Excel file for testing."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

SAMPLE_DATA = [
    ["PROJECT_ID", "FIRST_NAME", "MIDDLE_NAME", "LAST_NAME", "ADDRESS_LINE_1", "ADDRESS_LINE_2", "CITY", "STATE_CODE"],
    ["BI-001", "Valentin", "", "Fuster", "One Gustave L. Levy Place", "", "New York", "NY"],
    ["BI-002", "Ashish", "K", "Jha", "Brown University", "Box G-PH", "Providence", "RI"],
    ["BI-003", "Sanjay", "", "Gupta", "Emory University Hospital", "", "Atlanta", "GA"],
    ["BI-004", "Rochelle", "P", "Walensky", "Massachusetts General Hospital", "", "Boston", "MA"],
    ["BI-005", "Mikhail", "", "Varshavski", "Overlook Medical Center", "", "Summit", "NJ"],
]


def create_sample_excel(filepath: str = "sample_data/sample_hcp.xlsx") -> None:
    """Create a sample HCP Excel file with realistic test data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "HCP Records"

    # Header styling
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(bold=True, color="ffffff", size=11)

    for row_idx, row_data in enumerate(SAMPLE_DATA):
        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)
            if row_idx == 0:
                cell.fill = header_fill
                cell.font = header_font

    # Auto-width columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col) + 2
        ws.column_dimensions[col[0].column_letter].width = max_length

    wb.save(filepath)
    print(f"✅ Sample Excel created: {filepath}")


if __name__ == "__main__":
    create_sample_excel()
