"""Excel file ingestion — parse uploaded .xlsx / .xls into HCPInput records."""

from __future__ import annotations

from io import BytesIO
from typing import BinaryIO

from openpyxl import load_workbook

from hcp_crawler.models.schemas import HCPInput
from hcp_crawler.utils.logger import get_logger

logger = get_logger(__name__)

# Canonical column names (case-insensitive matching)
_EXPECTED_COLUMNS = {
    "PROJECT_ID",
    "FIRST_NAME",
    "MIDDLE_NAME",
    "LAST_NAME",
    "ADDRESS_LINE_1",
    "ADDRESS_LINE_2",
    "CITY",
    "STATE_CODE",
}


def _normalise_header(header: str) -> str:
    """Normalise a header string for matching."""
    return header.strip().upper().replace(" ", "_")


def parse_excel(file: BinaryIO | bytes, filename: str = "upload.xlsx") -> list[HCPInput]:
    """
    Parse an Excel file and return a list of validated HCPInput records.

    Raises ValueError if the file is unreadable or missing the PROJECT_ID column.
    """
    if isinstance(file, bytes):
        file = BytesIO(file)

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Could not read Excel file '{filename}': {exc}") from exc

    ws = wb.active
    if ws is None:
        raise ValueError("The Excel workbook has no active sheet.")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("The Excel file must have a header row and at least one data row.")

    # ── Map headers ──────────────────────────────────────────────────
    raw_headers = [str(h).strip() if h else "" for h in rows[0]]
    normalised = [_normalise_header(h) for h in raw_headers]

    if "PROJECT_ID" not in normalised:
        raise ValueError(
            "Missing required column 'PROJECT_ID'. "
            f"Found columns: {raw_headers}"
        )

    col_map: dict[str, int] = {}
    for idx, name in enumerate(normalised):
        if name in _EXPECTED_COLUMNS:
            col_map[name] = idx

    # ── Parse rows ───────────────────────────────────────────────────
    records: list[HCPInput] = []
    skipped = 0

    for row_num, row in enumerate(rows[1:], start=2):
        project_id_idx = col_map.get("PROJECT_ID")
        if project_id_idx is None:
            continue

        project_id = str(row[project_id_idx]).strip() if row[project_id_idx] else ""
        if not project_id:
            skipped += 1
            continue

        def _cell(col_name: str) -> str:
            idx = col_map.get(col_name)
            if idx is None or idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        records.append(
            HCPInput(
                project_id=project_id,
                first_name=_cell("FIRST_NAME"),
                middle_name=_cell("MIDDLE_NAME"),
                last_name=_cell("LAST_NAME"),
                address_line_1=_cell("ADDRESS_LINE_1"),
                address_line_2=_cell("ADDRESS_LINE_2"),
                city=_cell("CITY"),
                state_code=_cell("STATE_CODE"),
            )
        )

    wb.close()

    logger.info(
        "excel_parsed",
        filename=filename,
        total_rows=len(rows) - 1,
        valid_records=len(records),
        skipped=skipped,
    )
    return records
